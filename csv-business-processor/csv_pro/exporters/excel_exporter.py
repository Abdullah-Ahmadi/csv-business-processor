from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .base_exporter import BaseExporter
from utils.formatters import format_currency
from utils.file_utils import generate_filename
from utils.logger import get_logger


class ExcelExporter(BaseExporter):
    """
    Export processor insights to Excel format.
    """

    def __init__(self):
        self.styles = self._define_styles()
        self.logger = get_logger(self.__class__.__name__)

    def export(self, processor, output_path=None):
        self._validate_processor(processor)

        # Generate filename if not provided
        if output_path is None:
            output_path = generate_filename(processor, ".xlsx")

        if not output_path.endswith(".xlsx"):
            output_path += ".xlsx"

        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets based on processor type
        processor_type = processor.__class__.__name__

        if "Ecommerce" in processor_type:
            self._create_ecommerce_sheets(wb, processor)
        elif "Inventory" in processor_type:
            self._create_inventory_sheets(wb, processor)
        elif "Finance" in processor_type:
            self._create_finance_sheets(wb, processor)
        else:
            self._create_generic_sheets(wb, processor)

        # Save workbook
        try:
            wb.save(output_path)
            print(f"✅ Excel report saved to: {output_path}")
            return output_path

        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
            raise

    def _create_ecommerce_sheets(self, wb, processor):
        """Create sheets for e-commerce data."""

        # Executive Summary Sheet
        ws_summary = wb.create_sheet(title="Executive Summary")
        self._add_executive_summary(ws_summary, processor)

        # Sales Analysis Sheet
        ws_sales = wb.create_sheet(title="Sales Analysis")
        self._add_sales_analysis(ws_sales, processor)

        # Product Performance Sheet
        if "top_products" in processor.insights:
            ws_products = wb.create_sheet(title="Product Performance")
            self._add_product_performance(ws_products, processor)

        # Raw Data Sheet
        if processor.data is not None:
            ws_raw = wb.create_sheet(title="Raw Data")
            self._add_raw_data(ws_raw, processor.data)

    def _create_inventory_sheets(self, wb, processor):
        """Create sheets for inventory data."""
        # Inventory Health Sheet
        ws_health = wb.create_sheet(title="Inventory Health")
        self._add_inventory_health(ws_health, processor)

        # Restock Recommendations
        if "restock_recommendations" in processor.insights:
            ws_restock = wb.create_sheet(title="Restock Recommendations")
            self._add_restock_recommendations(ws_restock, processor)

        # Raw Data Sheet
        if processor.data is not None:
            ws_raw = wb.create_sheet(title="Raw Data")
            self._add_raw_data(ws_raw, processor.data)

    def _create_finance_sheets(self, wb, processor):
        """Create sheets for financial data."""
        # Spending Overview Sheet
        ws_overview = wb.create_sheet(title="Spending Overview")
        self._add_spending_overview(ws_overview, processor)

        # Category Analysis Sheet
        if "spending_by_category" in processor.insights:
            ws_categories = wb.create_sheet(title="Category Analysis")
            self._add_category_analysis(ws_categories, processor)  # This was missing!

        # Raw Data Sheet
        if processor.data is not None:
            ws_raw = wb.create_sheet(title="Raw Data")
            self._add_raw_data(ws_raw, processor.data)

    def _create_generic_sheets(self, wb, processor):
        """Create generic sheets for any processor type."""
        # Insights Sheet
        ws_insights = wb.create_sheet(title="Insights")
        self._add_generic_insights(ws_insights, processor)

        # Alerts Sheet
        if processor.alerts:
            ws_alerts = wb.create_sheet(title="Alerts")
            self._add_alerts(ws_alerts, processor.alerts)

        # Raw Data Sheet
        if processor.data is not None:
            ws_raw = wb.create_sheet(title="Raw Data")
            self._add_raw_data(ws_raw, processor.data)

    def _add_executive_summary(self, ws, processor):
        """Add executive summary to worksheet."""
        ws.title = "Executive Summary"
        ws.merge_cells("A1:D1")
        ws["A1"] = (
            f"{processor.__class__.__name__.replace('Processor', '')} Analysis Report"
        )
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Timestamp
        ws.merge_cells("A2:D2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # Summary metrics
        row = 4
        summary_metrics = ws.cell(row=row, column=1, value="📊 SUMMARY METRICS")
        summary_metrics.font = Font(bold=True, size=12)
        summary_metrics.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        summary = processor.get_processing_summary()
        for key, value in summary.items():
            ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(
                bold=True
            )
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Key insights
        row += 3
        key_insights = ws.cell(row=row, column=1, value="🔍 KEY INSIGHTS")
        key_insights.font = Font(bold=True, size=12)
        key_insights.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        for key, value in processor.insights.items():
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                display_key = key.replace("_", " ").title()
                display_value = (
                    format_currency(value)
                    if any(
                        term in key.lower()
                        for term in [
                            "revenue",
                            "amount",
                            "cost",
                            "price",
                            "spent",
                            "value",
                        ]
                    )
                    else value
                )
                ws.cell(row=row, column=1, value=display_key)
                ws.cell(row=row, column=2, value=display_value)
                row += 1

        # Adjust column widths
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _add_sales_analysis(self, ws, processor):
        """Add sales analysis data."""
        ws.title = "Sales Analysis"

        if "total_revenue" in processor.insights:
            metrics = [
                ("Total Revenue", processor.insights["total_revenue"]),
                ("Total Orders", processor.insights.get("total_orders", 0)),
                (
                    "Average Order Value",
                    processor.insights.get("average_order_value", 0),
                ),
                ("Total Units Sold", processor.insights.get("total_units_sold", 0)),
            ]

            row = 1
            ws.column_dimensions["A"].width = 20
            for label, value in metrics:
                ws["B" + str(row)].alignment = Alignment(horizontal="left")
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                if isinstance(value, (int, float)):
                    ws.cell(
                        row=row,
                        column=2,
                        value=(
                            format_currency(value)
                            if "revenue" in label.lower() or "value" in label.lower()
                            else value
                        ),
                    )
                else:
                    ws.cell(row=row, column=2, value=value)
                row += 1

    def _add_raw_data(self, ws, data):
        """Add raw data to worksheet."""
        if data is None or data.empty:
            return

        # Headers
        for col_idx, column in enumerate(data.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF")

        # Data
        for row_idx, row_data in enumerate(data.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _define_styles(self):
        """Define Excel styles."""
        return {
            "header": {
                "fill": PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                ),
                "font": Font(color="FFFFFF", bold=True),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "currency": {"number_format": '"$"#,##0.00'},
            "percentage": {"number_format": "0.00%"},
        }

    def _add_inventory_health(self, ws, processor):
        """Add inventory health analysis to worksheet."""
        ws.title = "Inventory Health"

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "Inventory Health Dashboard"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )

        # Timestamp
        ws.merge_cells("A2:F2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # Key metrics
        row = 4
        ws.cell(row=row, column=1, value="📊 KEY METRICS").font = Font(
            bold=True, size=12
        )
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        metrics = [
            ("Total Products", processor.insights.get("total_products", 0)),
            ("Total Stock Units", processor.insights.get("total_stock_units", 0)),
            (
                "Total Inventory Value",
                processor.insights.get("total_inventory_value", 0),
            ),
            ("Low Stock Items", processor.insights.get("low_stock_count", 0)),
            ("Overstock Items", processor.insights.get("overstock_count", 0)),
        ]

        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            if "Value" in label:
                ws.cell(row=row, column=2, value=value).number_format = '"$"#,##0.00'
            else:
                ws.cell(row=row, column=2, value=value)
            row += 1

        # Low stock items table
        if (
            "low_stock_items" in processor.insights
            and processor.insights["low_stock_items"]
        ):
            row += 2
            ws.cell(row=row, column=1, value="⚠️ LOW STOCK ITEMS").font = Font(
                bold=True, size=12, color="FF0000"
            )
            row += 1

            headers = ["Product", "Current Stock", "Reorder Level"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

            for item in processor.insights["low_stock_items"][:10]:  # Top 10
                row += 1
                ws.cell(row=row, column=1, value=item.get("product_name", "Unknown"))
                ws.cell(row=row, column=2, value=item.get("current_stock", 0))
                ws.cell(row=row, column=3, value=item.get("reorder_level", 0))

        # Overstock items table
        if (
            "overstock_items" in processor.insights
            and processor.insights["overstock_items"]
        ):
            row += 2
            ws.cell(row=row, column=1, value="📦 OVERSTOCK ITEMS").font = Font(
                bold=True, size=12, color="FF9900"
            )
            row += 1

            headers = ["Product", "Current Stock", "Reorder Level"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )

            for item in processor.insights["overstock_items"][:10]:  # Top 10
                row += 1
                ws.cell(row=row, column=1, value=item.get("product_name", "Unknown"))
                ws.cell(row=row, column=2, value=item.get("current_stock", 0))
                ws.cell(row=row, column=3, value=item.get("reorder_level", 0))

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _add_restock_recommendations(self, ws, processor):
        """Add restock recommendations to worksheet."""
        ws.title = "Restock Recommendations"

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "🔄 Restock Recommendations"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )

        row = 3
        headers = ["Product", "Order Quantity", "Estimated Cost", "Urgency"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)

        if "restock_recommendations" in processor.insights:
            for rec in processor.insights["restock_recommendations"]:
                row += 1
                ws.cell(row=row, column=1, value=rec.get("product", "Unknown"))
                ws.cell(row=row, column=2, value=rec.get("order_quantity", 0))
                ws.cell(
                    row=row, column=3, value=rec.get("estimated_cost", 0)
                ).number_format = '"$"#,##0.00'

                urgency = rec.get("urgency", "MEDIUM")
                cell = ws.cell(row=row, column=4, value=urgency)

                if urgency == "HIGH":
                    cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                elif urgency == "MEDIUM":
                    cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )

            # Total cost
            row += 2
            ws.cell(row=row, column=1, value="Total Restock Cost:").font = Font(
                bold=True
            )
            total_cost = sum(
                rec.get("estimated_cost", 0)
                for rec in processor.insights["restock_recommendations"]
            )
            ws.cell(row=row, column=2, value=total_cost).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=2).font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12

    def _add_product_performance(self, ws, processor):
        """Add product performance analysis to worksheet."""
        ws.title = "Product Performance"

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "🏆 Product Performance Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )

        # Check if we have product data
        if "top_products" not in processor.insights:
            ws.cell(
                row=3, column=1, value="No product performance data available."
            ).font = Font(italic=True)
            return

        top_products = processor.insights["top_products"]

        # Headers
        row = 3
        headers = ["Product", "Revenue", "Units Sold", "Orders", "Avg Order Value"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Product data
        row += 1
        if isinstance(top_products, dict):
            for product, stats in top_products.items():
                if isinstance(stats, dict):
                    ws.cell(row=row, column=1, value=product)

                    revenue = stats.get("revenue", stats.get("total_amount", 0))
                    ws.cell(row=row, column=2, value=revenue).number_format = (
                        '"$"#,##0.00'
                    )

                    units = stats.get("units_sold", stats.get("quantity", 0))
                    ws.cell(row=row, column=3, value=units)

                    orders = stats.get("order_count", stats.get("orders", 0))
                    ws.cell(row=row, column=4, value=orders)

                    avg_order = revenue / orders if orders > 0 else 0
                    ws.cell(row=row, column=5, value=avg_order).number_format = (
                        '"$"#,##0.00'
                    )

                    row += 1

        # Add totals row
        row += 1
        ws.cell(row=row, column=1, value="TOTALS / AVERAGES").font = Font(bold=True)

        # Calculate totals
        if isinstance(top_products, dict):
            total_revenue = sum(
                stats.get("revenue", stats.get("total_amount", 0))
                for stats in top_products.values()
                if isinstance(stats, dict)
            )
            total_units = sum(
                stats.get("units_sold", stats.get("quantity", 0))
                for stats in top_products.values()
                if isinstance(stats, dict)
            )
            total_orders = sum(
                stats.get("order_count", stats.get("orders", 0))
                for stats in top_products.values()
                if isinstance(stats, dict)
            )

            ws.cell(row=row, column=2, value=total_revenue).number_format = (
                '"$"#,##0.00'
            )
            ws.cell(row=row, column=3, value=total_units)
            ws.cell(row=row, column=4, value=total_orders)

            avg_all = total_revenue / total_orders if total_orders > 0 else 0
            ws.cell(row=row, column=5, value=avg_all).number_format = '"$"#,##0.00'

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15

    def _add_spending_overview(self, ws, processor):
        """Add spending overview to worksheet."""
        ws.title = "Spending Overview"

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "💰 Spending Overview & Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )

        # Timestamp
        ws.merge_cells("A2:E2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # Key metrics section
        row = 4
        ws.merge_cells(f"A{row}:B{row}")
        ws.cell(row=row, column=1, value="📊 KEY METRICS").font = Font(
            bold=True, size=12
        )
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        metrics = [
            ("Total Spent", processor.insights.get("total_spent", 0)),
            ("Transaction Count", processor.insights.get("transaction_count", 0)),
            ("Average Transaction", processor.insights.get("average_transaction", 0)),
            (
                "Top Category",
                processor.insights.get("top_category", "N/A").capitalize(),
            ),
            ("Top Category Amount", processor.insights.get("top_category_amount", 0)),
        ]

        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            if isinstance(value, (int, float)) and any(
                term in label.lower() for term in ["spent", "amount", "transaction"]
            ):
                ws.cell(row=row, column=2, value=value).number_format = '"$"#,##0.00'
            else:
                ws.cell(row=row, column=2, value=value)
            row += 1

        # Category breakdown
        if "spending_by_category" in processor.insights:
            row += 2
            ws.merge_cells(f"A{row}:B{row}")
            ws.cell(row=row, column=1, value="📈 SPENDING BY CATEGORY").font = Font(
                bold=True, size=12
            )
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
            )

            row += 1

            # Headers
            headers = ["Category", "Amount", "% of Total"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

            # Category data
            row += 1
            total_spent = processor.insights.get("total_spent", 0)
            for category, amount in processor.insights["spending_by_category"].items():
                ws.cell(row=row, column=1, value=category.capitalize())
                ws.cell(row=row, column=2, value=amount).number_format = '"$"#,##0.00'

                percentage = (amount / total_spent * 100) if total_spent > 0 else 0
                ws.cell(row=row, column=3, value=percentage).number_format = "0.00%"

                # Highlight high spending categories
                if percentage > 30:
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                row += 1

            # Total row
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=2, value=total_spent).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=3, value=1.0).number_format = "0.00%"

        # Monthly spending if available
        if "monthly_spending" in processor.insights:
            row += 3
            ws.merge_cells(f"A{row}:B{row}")
            ws.cell(row=row, column=1, value="📅 MONTHLY SPENDING TREND").font = Font(
                bold=True, size=12
            )
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
            )
            row += 1

            # Headers
            month = ws.cell(row=row, column=1, value="Month")
            month.font = Font(bold=True, color="FFFFFF")
            month.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            amount = ws.cell(row=row, column=2, value="Amount")
            amount.font = Font(bold=True, color="FFFFFF")
            amount.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            change = ws.cell(row=row, column=3, value="Change")
            change.font = Font(bold=True, color="FFFFFF")
            change.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            row += 1
            previous = None

            for month, amount in processor.insights["monthly_spending"].items():
                m = datetime.strptime(month, "%Y-%m")
                ws.cell(row=row, column=1, value=m.strftime("%b-%y"))
                ws.cell(row=row, column=2, value=amount).number_format = '"$"#,##0.00'

                if previous:
                    change = ((amount - previous) / previous) * 100
                    cell = ws.cell(row=row, column=3, value=change / 100)
                    cell.number_format = "0.00%"

                    if change > 20:
                        cell.font = Font(color="FF0000", bold=True)
                    elif change < -20:
                        cell.font = Font(color="00B050", bold=True)
                else:
                    ws.cell(row=row, column=3, value="N/A")

                previous = amount
                row += 1

        # Payment method analysis
        if "payment_method_stats" in processor.insights:
            row += 2
            ws.merge_cells(f"A{row}:B{row}")
            payment_method_cell = ws.cell(
                row=row, column=1, value="💳 PAYMENT METHOD ANALYSIS"
            )
            payment_method_cell.font = Font(bold=True, size=12)
            payment_method_cell.fill = PatternFill(
                start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
            )
            row += 1

            # Headers
            headers = ["Method", "Total", "Transactions", "Average"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

            row += 1
            for method, stats in processor.insights["payment_method_stats"].items():
                ws.cell(row=row, column=1, value=method)
                ws.cell(
                    row=row, column=2, value=stats.get("total", 0)
                ).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=3, value=stats.get("count", 0))
                ws.cell(
                    row=row, column=4, value=stats.get("average", 0)
                ).number_format = '"$"#,##0.00'
                row += 1

        # Savings opportunities if available
        if "savings_opportunities" in processor.insights:
            row += 2
            ws.merge_cells(f"A{row}:B{row}")
            cell = ws.cell(row=row, column=1, value="💰 SAVINGS OPPORTUNITIES")
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(
                start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
            )
            row += 1

            # Headers
            headers = ["Category", "Current Spending", "Potential Savings", "Action"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="00B050", end_color="00B050", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

            row += 1
            for opp in processor.insights["savings_opportunities"]:
                ws.cell(row=row, column=1, value=opp.get("category", "").capitalize())
                ws.cell(
                    row=row, column=2, value=opp.get("current_spending", 0)
                ).number_format = '"$"#,##0.00'
                ws.cell(
                    row=row, column=3, value=opp.get("suggested_reduction", 0)
                ).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=4, value=opp.get("reason", ""))
                row += 1

        # Large transactions if available
        if "large_transactions" in processor.insights:
            row += 2
            ws.merge_cells(f"A{row}:B{row}")
            cell = ws.cell(row=row, column=1, value="⚠️ LARGE TRANSACTIONS")
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(
                start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
            )
            row += 1

            # Headers
            headers = ["Date", "Description", "Amount", "Category"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

            row += 1
            for trans in processor.insights["large_transactions"][:10]:
                ws.cell(
                    row=row,
                    column=1,
                    value=datetime.strftime(trans.get("date", ""), "%Y-%m-%d"),
                )
                ws.cell(row=row, column=2, value=str(trans.get("description", ""))[:30])
                ws.cell(
                    row=row, column=3, value=trans.get("amount", 0)
                ).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=4, value=trans.get("category", "").capitalize())
                row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _add_category_analysis(self, ws, processor):
        """Add category analysis to worksheet."""
        ws.title = "Category Analysis"
        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "📊 Category Spending Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )

        row = 4

        if "spending_by_category" not in processor.insights:
            ws.cell(row=row, column=1, value="No category data available.").font = Font(
                italic=True
            )
            return

        # Category breakdown table
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value="📈 SPENDING BY CATEGORY")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        # Headers
        headers = [
            "Category",
            "Amount",
            "% of Total",
            "Transaction Count",
            "Avg per Transaction",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        row += 1
        total_spent = processor.insights.get("total_spent", 0)

        # Get transaction counts per category if available
        category_counts = {}
        if processor.data is not None and "category" in processor.data.columns:
            category_counts = processor.data.groupby("category").size().to_dict()

        for category, amount in processor.insights["spending_by_category"].items():
            ws.cell(row=row, column=1, value=category.capitalize())

            amount_cell = ws.cell(row=row, column=2, value=amount)
            amount_cell.number_format = '"$"#,##0.00'

            percentage = (amount / total_spent * 100) if total_spent > 0 else 0
            pct_cell = ws.cell(row=row, column=3, value=percentage / 100)
            pct_cell.number_format = "0.00%"

            count = category_counts.get(category, 0)
            ws.cell(row=row, column=4, value=count)

            avg = amount / count if count > 0 else 0
            avg_cell = ws.cell(row=row, column=5, value=avg)
            avg_cell.number_format = '"$"#,##0.00'

            if percentage > 30:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
            elif percentage < 5:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )

            row += 1

        row += 1
        ws.cell(row=row, column=1, value="TOTAL / AVERAGE").font = Font(bold=True)

        total_cell = ws.cell(row=row, column=2, value=total_spent)
        total_cell.number_format = '"$"#,##0.00'
        total_cell.font = Font(bold=True)

        pct_cell = ws.cell(row=row, column=3, value=1.0)
        pct_cell.number_format = "0.00%"
        pct_cell.font = Font(bold=True)

        total_count = sum(category_counts.values())
        ws.cell(row=row, column=4, value=total_count).font = Font(bold=True)

        avg_total = total_spent / total_count if total_count > 0 else 0
        avg_cell = ws.cell(row=row, column=5, value=avg_total)
        avg_cell.number_format = '"$"#,##0.00'
        avg_cell.font = Font(bold=True)

        # Category percentages chart data for visual reference
        row += 3
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value="📊 CATEGORY DISTRIBUTION")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        cat_cell = ws.cell(row=row, column=1, value="Category")
        cat_cell.font = Font(bold=True, color="FFFFFF")
        cat_cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        pct_cell = ws.cell(row=row, column=2, value="Percentage")
        pct_cell.font = Font(bold=True, color="FFFFFF")
        pct_cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        vsl_cell = ws.cell(row=row, column=3, value="Visual")
        vsl_cell.font = Font(bold=True, color="FFFFFF")
        vsl_cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        row += 1
        for category, amount in processor.insights["spending_by_category"].items():
            percentage = (amount / total_spent * 100) if total_spent > 0 else 0
            ws.cell(row=row, column=1, value=category.capitalize())
            pct_cell = ws.cell(row=row, column=2, value=percentage / 100)
            pct_cell.number_format = "0.00%"

            # Create simple bar chart using characters
            bar_length = int(
                percentage / 2
            )  # Scale for display (max 50 chars for 100%)
            ws.cell(row=row, column=3, value="█" * bar_length)
            row += 1

        # Category insights
        row += 2
        ws.merge_cells(f"A{row}:B{row}")
        cat_cell = ws.cell(row=row, column=1, value="🔍 CATEGORY INSIGHTS")
        cat_cell.font = Font(bold=True, size=12)
        cat_cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        # Find highest category
        if processor.insights["spending_by_category"]:
            top_category = max(
                processor.insights["spending_by_category"],
                key=processor.insights["spending_by_category"].get,
            )
            top_amount = processor.insights["spending_by_category"][top_category]
            top_pct = (top_amount / total_spent * 100) if total_spent > 0 else 0

            ws.cell(row=row, column=1, value="Highest spending category")
            ws.cell(row=row, column=2, value=f"{top_category.capitalize()}")
            ws.cell(row=row, column=3, value=f"${top_amount:,.2f} ({top_pct:.1f}%)")
            row += 1

        # Find lowest category (excluding zeros)
        non_zero = {
            k: v for k, v in processor.insights["spending_by_category"].items() if v > 0
        }
        if non_zero:
            low_category = min(non_zero, key=non_zero.get)
            low_amount = non_zero[low_category]
            low_pct = (low_amount / total_spent * 100) if total_spent > 0 else 0

            ws.cell(row=row, column=1, value="Lowest spending category")
            ws.cell(row=row, column=2, value=f"{low_category.capitalize()}")
            ws.cell(row=row, column=3, value=f"${low_amount:,.2f} ({low_pct:.1f}%)")
            row += 1

        # Average spending per category
        avg_category = (
            total_spent / len(processor.insights["spending_by_category"])
            if processor.insights["spending_by_category"]
            else 0
        )
        ws.cell(row=row, column=1, value="Average per category")
        ws.cell(row=row, column=3, value=f"${avg_category:,.2f}")
        row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20

    def _add_generic_insights(self, ws, processor):
        """Add generic insights to worksheet for any processor type."""
        ws.title = "Insights"

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "📋 Analysis Insights"
        ws["A1"].font = Font(size=16, bold=True, color="366092")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(
            start_color="FFFD55", end_color="FFFD55", fill_type="solid"
        )

        row = 3

        # Processor info
        cell = ws.cell(row=row, column=1, value="📊 PROCESSOR INFORMATION")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )

        row += 1

        insight_summary = processor.get_processing_summary()
        info_data = [
            (
                "Processor Type",
                f"{insight_summary.get('processing_mode', 'Unknown')} Processor",
            ),
            ("Rows Processed", insight_summary.get("rows_processed", "N/A")),
            ("Columns Found", insight_summary.get("columns", "N/A")),
            ("Input File", getattr(processor, "input_file", "Unknown")),
        ]

        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Key Insights Section
        row += 2
        cell = ws.cell(row=row, column=1, value="🔍 KEY INSIGHTS")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        if processor.insights:
            insight_count = 0
            for key, value in processor.insights.items():
                # Skip complex structures
                if isinstance(value, (dict, list)) and len(str(value)) > 100:
                    continue

                # Format the insight
                display_key = key.replace("_", " ").title()

                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    # Format financial metrics
                    if any(
                        term in key.lower()
                        for term in [
                            "revenue",
                            "amount",
                            "cost",
                            "price",
                            "spent",
                            "value",
                            "total",
                        ]
                    ):
                        display_value = f"${value:,.2f}"
                    elif any(
                        term in key.lower()
                        for term in ["percentage", "rate", "growth", "margin"]
                    ):
                        display_value = f"{value:.1f}%"
                    else:
                        display_value = f"{value:,}"
                else:
                    display_value = str(value)[:50]  # Truncate long strings

                ws.cell(row=row, column=1, value=f"• {display_key}:").font = Font(
                    bold=True
                )
                ws.cell(row=row, column=2, value=display_value)
                row += 1
                insight_count += 1

                # Limit to 20 insights to avoid overwhelming
                if insight_count >= 20:
                    ws.cell(
                        row=row,
                        column=1,
                        value="• ... and more insights (see JSON export for details)",
                    ).font = Font(italic=True)
                    row += 1
                    break
        else:
            ws.cell(row=row, column=1, value="No insights generated.").font = Font(
                italic=True
            )
            row += 1

        # Summary Statistics
        row += 2
        cell = ws.cell(row=row, column=1, value="📈 SUMMARY STATISTICS")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )

        row += 1

        if processor.data is not None and not processor.data.empty:
            numeric_cols = processor.data.select_dtypes(include=["number"]).columns

            for col in numeric_cols[:5]:  # Limit to first 5 numeric columns
                stats = [
                    (f"{col} - Total", processor.data[col].sum()),
                    (f"{col} - Average", processor.data[col].mean()),
                    (f"{col} - Min", processor.data[col].min()),
                    (f"{col} - Max", processor.data[col].max()),
                ]

                for label, stat_value in stats:
                    ws.cell(row=row, column=1, value=f"• {label}:")

                    if any(
                        term in col.lower()
                        for term in [
                            "price",
                            "cost",
                            "amount",
                            "revenue",
                            "spent",
                            "value",
                        ]
                    ):
                        ws.cell(row=row, column=2, value=f"${stat_value:,.2f}")
                    else:
                        ws.cell(row=row, column=2, value=f"{stat_value:,.2f}")
                    row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _add_alerts(self, ws, alerts):
        """Add alerts to worksheet."""
        ws.title = "Alerts"

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "🚨 Alerts & Recommendations"
        ws["A1"].font = Font(size=16, bold=True, color="FF0000")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )

        row = 3

        if not alerts:
            ws.cell(
                row=row,
                column=1,
                value="✅ No alerts generated. Everything looks good!",
            ).font = Font(color="00B050", bold=True)
            return

        # Summary header
        ws.cell(row=row, column=1, value="Priority").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Alert").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Category").font = Font(bold=True)
        ws.cell(row=row, column=4, value="Action Required").font = Font(bold=True)

        # Style headers
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Categorize alerts by severity
        critical_count = 0
        warning_count = 0
        info_count = 0

        for alert in alerts:
            # Determine priority based on content
            alert_text = str(alert)
            priority = "INFO"

            if any(
                term in alert_text.lower()
                for term in ["critical", "urgent", "out of stock", "⚠️"]
            ):
                priority = "CRITICAL"
                critical_count += 1
            elif any(
                term in alert_text.lower()
                for term in ["warning", "low stock", "high", "alert"]
            ):
                priority = "WARNING"
                warning_count += 1
            else:
                info_count += 1

            # Determine category
            category = "General"
            if any(
                term in alert_text.lower() for term in ["stock", "inventory", "restock"]
            ):
                category = "Inventory"
            elif any(
                term in alert_text.lower() for term in ["sales", "revenue", "product"]
            ):
                category = "Sales"
            elif any(
                term in alert_text.lower()
                for term in ["spending", "budget", "expense", "cost"]
            ):
                category = "Finance"
            elif any(term in alert_text.lower() for term in ["rating", "review"]):
                category = "Customer"

            # Priority cell
            priority_cell = ws.cell(row=row, column=1, value=priority)
            if priority == "CRITICAL":
                priority_cell.font = Font(bold=True, color="FFFFFF")
                priority_cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
            elif priority == "WARNING":
                priority_cell.font = Font(bold=True, color="000000")
                priority_cell.fill = PatternFill(
                    start_color="FFC000", end_color="FFC000", fill_type="solid"
                )
            else:
                priority_cell.font = Font(bold=True, color="000000")
                priority_cell.fill = PatternFill(
                    start_color="00B0F0", end_color="00B0F0", fill_type="solid"
                )

            # Alert text
            ws.cell(row=row, column=2, value=alert_text[:100])  # Truncate if too long

            # Category
            ws.cell(row=row, column=3, value=category)

            # Suggested action (simplified)
            action = self._suggest_action(alert_text, category)
            ws.cell(row=row, column=4, value=action)

            row += 1

        # Summary statistics
        row += 2
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value="📊 ALERTS SUMMARY")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )

        row += 1

        summary_data = [
            ("Total Alerts", len(alerts)),
            ("Critical", critical_count),
            ("Warnings", warning_count),
            ("Informational", info_count),
        ]

        for label, count in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Recommendations section
        row += 2
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value="✅ RECOMMENDED ACTIONS")
        cell.font = Font(bold=True, size=12, color="00B050")
        cell.fill = PatternFill(
            start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
        )
        row += 1

        recommendations = self._generate_recommendations(alerts)
        for i, rec in enumerate(recommendations[:5], 1):
            ws.cell(row=row, column=1, value=f"{i}.")
            ws.cell(row=row, column=2, value=rec)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40

    def _suggest_action(self, alert_text, category):
        """Suggest an action based on alert text and category."""
        alert_lower = alert_text.lower()

        if category == "Inventory":
            if "out of stock" in alert_lower:
                return "Place urgent order immediately"
            elif "low stock" in alert_lower:
                return "Review reorder levels and place order"
            elif "overstock" in alert_lower:
                return "Consider running a promotion to clear excess"

        elif category == "Sales":
            if "decline" in alert_lower:
                return "Analyze competitor pricing and marketing"
            elif "increase" in alert_lower:
                return "Ensure sufficient stock for continued growth"

        elif category == "Finance":
            if "spending" in alert_lower:
                return "Review category budget and identify savings"
            elif "large transaction" in alert_lower:
                return "Verify transaction legitimacy and budget impact"

        elif category == "Customer":
            if "rating" in alert_lower:
                return "Investigate product quality or customer service"

        return "Review and investigate further"

    def _generate_recommendations(self, alerts):
        """Generate actionable recommendations from alerts."""
        recommendations = []

        # Combine alerts into actionable items
        if any("stock" in str(a).lower() for a in alerts):
            recommendations.append("Run inventory audit and update reorder levels")

        if any("spending" in str(a).lower() for a in alerts):
            recommendations.append(
                "Review budget allocation and identify cost-saving opportunities"
            )

        if any(
            "sales" in str(a).lower() or "revenue" in str(a).lower() for a in alerts
        ):
            recommendations.append("Analyze sales trends and adjust marketing strategy")

        if any("rating" in str(a).lower() for a in alerts):
            recommendations.append(
                "Follow up with customers and address product quality issues"
            )

        # Add generic recommendations if needed
        if len(recommendations) < 3:
            recommendations.append(
                "Schedule regular data reviews to catch issues early"
            )
            recommendations.append(
                "Export data to different formats for deeper analysis"
            )

        return recommendations
